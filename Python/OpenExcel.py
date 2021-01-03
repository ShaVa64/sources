from openpyxl import Workbook, load_workbook
import pandas as pd

# ,strArray ,dd
# Read EXCEL file contents to an array


def ReadExcelSheet (strFile,strSheet):
    data = pd.ExcelFile(strFile)
    sheet = data.parse(strSheet)
    return True,sheet


# Ready to manage array of lists or whatever works ...
def ReadExcelWbToArray (wb,arrRules):
    ws = wb.active
    print (str(wb.sheetnames))
    nRow=0
    for row in ws.rows:
        nRow += 1
        nCell=0
        for cell in row:
            nCell += 1
            ## Replace here :: 
            arrRules[nRow,nCell]=cell.value
            print(str(cell.value))
    return True

def LoadExcelWb( strExcelFineName):
    wb = load_workbook(filename = strExcelFineName, read_only=True)
    print (str(wb.sheetnames))
    ws = wb.active
    nRows= ws.max_row
    nCells= ws.max_column
    return wb,nRows,nCells


def nam_ReadExcelWbToArray ( wb ,arrRules):
    ws = wb.active
    print (str(wb.sheetnames))
    nRow=0
    for row in ws.rows:
        nRow += 1
        nCell=0
        for cell in row:
            nCell += 1
            arrRules[nRow,nCell]=cell.value
            print(str(cell.value))
    return True

def nam_ReadExcelFileToArray( strExcelFineName ,arrRules):
    wb = load_workbook(filename = strExcelFineName, read_only=True)
    print (str(wb.sheetnames))
    ws = wb.active
    nRow=0
    for row in ws.rows:
        nRow += 1
        nCell=0
        for cell in row:
            nCell += 1
            arrRules[nRow,nCell]=cell.value
            print(str(cell.value))
    return True

""" 
# Comment long stuff :

## Sample code from https://openpyxl.readthedocs.io/en/stable/
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

def convert_coordinates(coordinates):
    # list[x, y] -> string['A1']
    pass """